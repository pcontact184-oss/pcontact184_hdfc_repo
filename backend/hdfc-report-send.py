import json
import os
import re
import boto3
from io import BytesIO
from datetime import datetime, timedelta, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ==========================================================
# ENVIRONMENT VARIABLES
# ==========================================================

SES_REGION = os.environ.get("SES_REGION", "ap-south-1")
API_KEY = os.environ.get("API_KEY", "")
SENDER_EMAIL = os.environ.get("SENDER_EMAIL", "")
TIME_PERIOD = os.environ.get("TIME_PERIOD", "30")
REGIONS = os.environ.get("REGIONS", "ap-south-1")

# Optional: comma-separated fallback recipients from env if needed
DEFAULT_TO_EMAILS = os.environ.get("DEFAULT_TO_EMAILS", "")

# ==========================================================
# AWS CLIENTS
# ==========================================================

ses = boto3.client("ses", region_name=SES_REGION)


# ==========================================================
# COMMON HELPERS
# ==========================================================

def cors_headers():
    return {
        "Access-Control-Allow-Origin": "*",
        "Access-Control-Allow-Headers": "*",
        "Access-Control-Allow-Methods": "POST,OPTIONS",
        "Content-Type": "application/json"
    }


def resp(status, body):
    return {
        "statusCode": status,
        "headers": cors_headers(),
        "body": json.dumps(body, default=str)
    }


def get_route_path(event):
    raw_path = event.get("rawPath")
    if raw_path:
        return raw_path

    return (
        event.get("requestContext", {})
        .get("http", {})
        .get("path", "")
    )


def get_headers_lower(event):
    headers = event.get("headers") or {}
    return {str(k).lower(): v for k, v in headers.items()}


def parse_body(event):
    body = event.get("body")
    if not body:
        return {}

    # HTTP API can send base64-encoded body, but only decode if explicitly marked
    if event.get("isBase64Encoded"):
        try:
            import base64
            body = base64.b64decode(body).decode("utf-8")
        except Exception as e:
            print(f"Failed to decode base64 body: {e}")
            return {}

    if isinstance(body, dict):
        return body

    try:
        return json.loads(body)
    except Exception as e:
        print(f"Failed to parse JSON body: {e}")
        return {}


def sanitize_sheet_name(name, prefix="Sheet"):
    name = str(name or prefix)
    invalid_chars = r'[\\/*?:\[\]]'
    name = re.sub(invalid_chars, "_", name)
    name = name[:31]
    return name if name else prefix


def parse_days(body):
    raw_days = body.get("days", TIME_PERIOD)
    try:
        days = int(raw_days)
        if days <= 0:
            return int(TIME_PERIOD)
        return days
    except Exception:
        return int(TIME_PERIOD)


def parse_emails_from_body(body):
    raw = body.get("toEmails") or body.get("emails") or DEFAULT_TO_EMAILS or ""

    if isinstance(raw, list):
        emails = [str(x).strip().lower() for x in raw if str(x).strip()]
    else:
        emails = [e.strip().lower() for e in str(raw).split(",") if e.strip()]

    # de-duplicate while preserving order
    seen = set()
    unique = []
    for email in emails:
        if email not in seen:
            seen.add(email)
            unique.append(email)

    return unique


def parse_regions_from_body(body):
    raw = body.get("regions")

    if isinstance(raw, list):
        regions = [str(r).strip() for r in raw if str(r).strip()]
    elif isinstance(raw, str):
        regions = [r.strip() for r in raw.split(",") if r.strip()]
    else:
        regions = [r.strip() for r in REGIONS.split(",") if r.strip()]

    # de-duplicate while preserving order
    seen = set()
    unique = []
    for region in regions:
        if region not in seen:
            seen.add(region)
            unique.append(region)

    return unique


def is_valid_email(email):
    pattern = r"^[^@\s]+@[^@\s]+\.[^@\s]+$"
    return bool(re.match(pattern, email or ""))


def workbook_to_bytes(wb):
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# ==========================================================
# SES HELPERS
# ==========================================================

def get_verified_emails():
    identities = []
    token = None

    while True:
        params = {"IdentityType": "EmailAddress", "MaxItems": 1000}
        if token:
            params["NextToken"] = token

        resp_ids = ses.list_identities(**params)
        identities.extend([x.lower() for x in resp_ids.get("Identities", [])])

        token = resp_ids.get("NextToken")
        if not token:
            break

    return set(identities)


def send_email(subject, html, attachments, recipients):
    msg = MIMEMultipart()
    msg["From"] = SENDER_EMAIL
    msg["To"] = ", ".join(recipients)
    msg["Subject"] = subject

    msg.attach(MIMEText(html, "html"))

    for name, data in attachments.items():
        part = MIMEApplication(data)
        part.add_header("Content-Disposition", "attachment", filename=name)
        msg.attach(part)

    send_resp = ses.send_raw_email(
        Source=SENDER_EMAIL,
        Destinations=recipients,
        RawMessage={"Data": msg.as_string()}
    )

    print("SES send_raw_email response:", send_resp)
    return send_resp


# ==========================================================
# EXCEL STYLING
# ==========================================================

def apply_excel_style(ws):
    header_fill = PatternFill("solid", fgColor="FFD966")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center
            cell.border = border

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 15), 60)


# ==========================================================
# GUARDDUTY
# ==========================================================

def get_guardduty_detector_id(guardduty_client):
    detectors = guardduty_client.list_detectors().get("DetectorIds", [])
    if not detectors:
        return None
    return detectors[0]


def get_guardduty_findings(guardduty_client, start, end):
    """
    Primary fetch using updatedAt filter.
    Fallback: if API behavior or filter issue causes empty result,
    fetch all finding IDs and filter locally by UpdatedAt / CreatedAt.
    """
    detector_id = get_guardduty_detector_id(guardduty_client)
    if not detector_id:
        print("No GuardDuty detector found in this region.")
        return []

    start_ts = int(start.timestamp())
    end_ts = int(end.timestamp())

    finding_ids = []
    token = None

    # Primary method: server-side filter by updatedAt
    try:
        while True:
            params = {
                "DetectorId": detector_id,
                "FindingCriteria": {
                    "Criterion": {
                        "updatedAt": {
                            "Gte": start_ts,
                            "Lte": end_ts
                        }
                    }
                },
                "MaxResults": 50
            }
            if token:
                params["NextToken"] = token

            page = guardduty_client.list_findings(**params)
            page_ids = page.get("FindingIds", [])
            finding_ids.extend(page_ids)

            token = page.get("NextToken")
            if not token:
                break

        print(f"GuardDuty primary filtered finding count: {len(finding_ids)}")
    except Exception as e:
        print(f"GuardDuty filtered list_findings failed: {e}")
        finding_ids = []

    # Fallback method: fetch all IDs if filtered query returns 0
    if not finding_ids:
        print("GuardDuty fallback activated: fetching all findings and filtering locally.")
        token = None
        all_ids = []

        while True:
            params = {
                "DetectorId": detector_id,
                "MaxResults": 50
            }
            if token:
                params["NextToken"] = token

            page = guardduty_client.list_findings(**params)
            all_ids.extend(page.get("FindingIds", []))

            token = page.get("NextToken")
            if not token:
                break

        print(f"GuardDuty fallback total finding IDs fetched: {len(all_ids)}")

        findings = []
        for i in range(0, len(all_ids), 50):
            batch = all_ids[i:i + 50]
            details = guardduty_client.get_findings(
                DetectorId=detector_id,
                FindingIds=batch
            )
            findings.extend(details.get("Findings", []))

        filtered_findings = []
        for f in findings:
            updated_at = f.get("UpdatedAt")
            created_at = f.get("CreatedAt")

            matched = False
            for ts_str in [updated_at, created_at]:
                if not ts_str:
                    continue
                try:
                    # GuardDuty usually returns ISO format with Z
                    dt = datetime.fromisoformat(ts_str.replace("Z", "+00:00"))
                    if start <= dt.astimezone(timezone.utc) <= end:
                        matched = True
                        break
                except Exception as e:
                    print(f"Failed to parse GuardDuty timestamp '{ts_str}': {e}")

            if matched:
                filtered_findings.append(f)

        print(f"GuardDuty fallback filtered findings count: {len(filtered_findings)}")
        return filtered_findings

    # If primary method worked, fetch details
    findings = []
    for i in range(0, len(finding_ids), 50):
        batch = finding_ids[i:i + 50]
        details = guardduty_client.get_findings(
            DetectorId=detector_id,
            FindingIds=batch
        )
        findings.extend(details.get("Findings", []))

    print(f"GuardDuty findings detail count: {len(findings)}")
    return findings


def build_guardduty_excel(findings_by_region):
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    headers = [
        "Id",
        "Title",
        "Type",
        "Severity",
        "Region",
        "ResourceType",
        "AccountId",
        "CreatedAt",
        "UpdatedAt",
        "Description"
    ]

    if not findings_by_region:
        ws = wb.create_sheet("GuardDuty Findings")
        ws.append(headers)
        ws.append(["No findings found", "", "", "", "", "", "", "", "", ""])
        apply_excel_style(ws)
        return wb

    total_rows = 0
    for region, findings in findings_by_region.items():
        ws = wb.create_sheet(sanitize_sheet_name(f"GD_{region}", "GD_Report"))
        ws.append(headers)

        if findings:
            for f in findings:
                ws.append([
                    f.get("Id", ""),
                    f.get("Title", ""),
                    f.get("Type", ""),
                    f.get("Severity", ""),
                    f.get("Region", region),
                    f.get("Resource", {}).get("ResourceType", ""),
                    f.get("AccountId", ""),
                    f.get("CreatedAt", ""),
                    f.get("UpdatedAt", ""),
                    f.get("Description", ""),
                ])
                total_rows += 1
        else:
            ws.append(["No findings found in this region", "", "", "", region, "", "", "", "", ""])

        apply_excel_style(ws)

    # Summary sheet
    summary = wb.create_sheet("Summary", 0)
    summary.append(["Region", "GuardDuty Findings Count"])
    for region, findings in findings_by_region.items():
        summary.append([region, len(findings)])
    summary.append(["Total", total_rows])
    apply_excel_style(summary)

    return wb


# ==========================================================
# AWS CONFIG
# ==========================================================

def is_config_recorder_available(config_client):
    try:
        rec = config_client.describe_configuration_recorders().get("ConfigurationRecorders", [])
        status = config_client.describe_configuration_recorder_status().get("ConfigurationRecordersStatus", [])
        print(f"Config recorders: {len(rec)}, statuses: {len(status)}")
        return len(rec) > 0
    except Exception as e:
        print(f"Failed checking Config recorder: {e}")
        return False


def get_config_changes(config_client, start, end):
    """
    Pull Config items from advanced query.
    If no recorder / no data / unsupported in region, returns [].
    """
    if not is_config_recorder_available(config_client):
        print("AWS Config recorder not available in this region.")
        return []

    results = []
    query = """
    SELECT
      resourceId,
      resourceName,
      resourceType,
      awsRegion,
      configurationItemCaptureTime,
      accountId,
      configurationItemStatus
    WHERE configurationItemCaptureTime >= '{}'
      AND configurationItemCaptureTime <= '{}'
    """.format(
        start.strftime("%Y-%m-%dT%H:%M:%SZ"),
        end.strftime("%Y-%m-%dT%H:%M:%SZ")
    )

    print("AWS Config query:", query)

    token = None
    while True:
        params = {
            "Expression": query,
            "Limit": 100
        }
        if token:
            params["NextToken"] = token

        resp_query = config_client.select_resource_config(**params)
        rows = resp_query.get("Results", [])
        print(f"AWS Config page returned rows: {len(rows)}")

        for row in rows:
            try:
                results.append(json.loads(row))
            except Exception as e:
                print(f"Failed to parse Config row: {e}, row={row}")

        token = resp_query.get("NextToken")
        if not token:
            break

    print(f"AWS Config total results: {len(results)}")
    return results


def build_config_excel(items_by_region):
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    headers = [
        "ResourceId",
        "ResourceName",
        "ResourceType",
        "Region",
        "CaptureTime",
        "AccountId",
        "Status"
    ]

    if not items_by_region:
        ws = wb.create_sheet("AWS Config Changes")
        ws.append(headers)
        ws.append(["No config changes found", "", "", "", "", "", ""])
        apply_excel_style(ws)
        return wb

    total_rows = 0
    for region, items in items_by_region.items():
        ws = wb.create_sheet(sanitize_sheet_name(f"CFG_{region}", "CFG_Report"))
        ws.append(headers)

        if items:
            for item in items:
                ws.append([
                    item.get("resourceId", ""),
                    item.get("resourceName", ""),
                    item.get("resourceType", ""),
                    item.get("awsRegion", region),
                    item.get("configurationItemCaptureTime", ""),
                    item.get("accountId", ""),
                    item.get("configurationItemStatus", ""),
                ])
                total_rows += 1
        else:
            ws.append(["No config changes found in this region", "", "", region, "", "", ""])

        apply_excel_style(ws)

    # Summary sheet
    summary = wb.create_sheet("Summary", 0)
    summary.append(["Region", "Config Items Count"])
    for region, items in items_by_region.items():
        summary.append([region, len(items)])
    summary.append(["Total", total_rows])
    apply_excel_style(summary)

    return wb


# ==========================================================
# REPORT SEND HANDLER
# ==========================================================

def report_send_handler(event):
    headers = get_headers_lower(event)
    request_key = headers.get("x-api-key")

    if request_key != API_KEY:
        return resp(403, {"message": "Unauthorized"})

    if not SENDER_EMAIL:
        return resp(500, {"message": "SENDER_EMAIL environment variable is missing"})

    body = parse_body(event)

    print("REPORT SEND RAW BODY:", event.get("body"))
    print("REPORT SEND PARSED BODY:", body)

    days = parse_days(body)
    emails = parse_emails_from_body(body)
    selected_regions = parse_regions_from_body(body)

    print("REPORT SEND PARSED emails:", emails)
    print("REPORT SEND PARSED regions:", selected_regions)
    print("REPORT SEND PARSED days:", days)

    if not emails:
        return resp(400, {"message": "No email provided"})

    invalid_emails = [e for e in emails if not is_valid_email(e)]
    if invalid_emails:
        return resp(400, {
            "message": "Invalid email format found",
            "invalidEmails": invalid_emails
        })

    if not selected_regions:
        selected_regions = [SES_REGION]

    verified = get_verified_emails()
    not_verified = [e for e in emails if e not in verified]
    if not_verified:
        return resp(400, {
            "message": "Some emails are not verified in SES",
            "notVerified": not_verified
        })

    end = datetime.now(timezone.utc)
    start = end - timedelta(days=days)

    gd_findings_by_region = {}
    cfg_items_by_region = {}
    region_status = {}

    total_gd = 0
    total_cfg = 0

    for region in selected_regions:
        region_status[region] = {
            "guardduty": {"status": "not-started", "count": 0, "error": None},
            "config": {"status": "not-started", "count": 0, "error": None}
        }

        try:
            print(f"Fetching GuardDuty data for region: {region}")
            guardduty_client = boto3.client("guardduty", region_name=region)
            gd_findings = get_guardduty_findings(guardduty_client, start, end)
            gd_findings_by_region[region] = gd_findings
            total_gd += len(gd_findings)

            region_status[region]["guardduty"]["status"] = "success"
            region_status[region]["guardduty"]["count"] = len(gd_findings)

            print(f"Region {region}: GuardDuty findings count = {len(gd_findings)}")

        except Exception as gd_err:
            print(f"GuardDuty fetch failed for {region}: {gd_err}")
            gd_findings_by_region[region] = []
            region_status[region]["guardduty"]["status"] = "failed"
            region_status[region]["guardduty"]["error"] = str(gd_err)

        try:
            print(f"Fetching AWS Config data for region: {region}")
            config_client = boto3.client("config", region_name=region)
            cfg_items = get_config_changes(config_client, start, end)
            cfg_items_by_region[region] = cfg_items
            total_cfg += len(cfg_items)

            region_status[region]["config"]["status"] = "success"
            region_status[region]["config"]["count"] = len(cfg_items)

            print(f"Region {region}: AWS Config items count = {len(cfg_items)}")

        except Exception as cfg_err:
            print(f"AWS Config fetch failed for {region}: {cfg_err}")
            cfg_items_by_region[region] = []
            region_status[region]["config"]["status"] = "failed"
            region_status[region]["config"]["error"] = str(cfg_err)

    gd_bytes = workbook_to_bytes(build_guardduty_excel(gd_findings_by_region))
    cfg_bytes = workbook_to_bytes(build_config_excel(cfg_items_by_region))

    subject = f"AWS Security Report | Last {days} Days"
    regions_line = ", ".join(selected_regions)

    region_rows = ""
    for region, info in region_status.items():
        region_rows += f"""
        <tr>
            <td>{region}</td>
            <td>{info['guardduty']['count']}</td>
            <td>{info['guardduty']['status']}</td>
            <td>{info['config']['count']}</td>
            <td>{info['config']['status']}</td>
        </tr>
        """

    html = f"""
    <html>
      <body>
        <h2>AWS Security Report</h2>
        <p><strong>Reporting Period:</strong> Last {days} days</p>
        <p><strong>Start Time (UTC):</strong> {start.strftime('%Y-%m-%d %H:%M:%S UTC')}</p>
        <p><strong>End Time (UTC):</strong> {end.strftime('%Y-%m-%d %H:%M:%S UTC')}</p>
        <p><strong>Regions:</strong> {regions_line}</p>
        <p><strong>Total GuardDuty Findings:</strong> {total_gd}</p>
        <p><strong>Total Config Items:</strong> {total_cfg}</p>
        <hr>
        <p><strong>Region-wise status:</strong></p>
        <table border="1" cellpadding="6" cellspacing="0" style="border-collapse: collapse;">
          <tr>
            <th>Region</th>
            <th>GuardDuty Count</th>
            <th>GuardDuty Status</th>
            <th>Config Count</th>
            <th>Config Status</th>
          </tr>
          {region_rows}
        </table>
        <br>
        <p>Please find attached:</p>
        <ul>
          <li>guardduty_report.xlsx</li>
          <li>aws_config_report.xlsx</li>
        </ul>
      </body>
    </html>
    """

    send_email(
        subject,
        html,
        {
            "guardduty_report.xlsx": gd_bytes,
            "aws_config_report.xlsx": cfg_bytes,
        },
        emails,
    )

    return resp(200, {
        "message": "Report sent successfully",
        "emails": emails,
        "regions": selected_regions,
        "days": days,
        "guarddutyCount": total_gd,
        "configCount": total_cfg,
        "regionStatus": region_status
    })


# ==========================================================
# MAIN LAMBDA HANDLER
# ==========================================================

def lambda_handler(event, context):
    method = event.get("requestContext", {}).get("http", {}).get("method", "")

    if method == "OPTIONS":
        return {
            "statusCode": 200,
            "headers": cors_headers(),
            "body": ""
        }

    path = get_route_path(event)

    print("ROUTE path:", path, "method:", method)
    print("EVENT:", json.dumps(event, default=str))

    if path.endswith("/report/send"):
        return report_send_handler(event)

    return resp(404, {"message": f"Route not found: {path}"})